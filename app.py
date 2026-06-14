import streamlit as st
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import json, io, re, datetime

st.set_page_config(
    page_title="CS 결과보고 AI 자동작성",
    page_icon="📊",
    layout="wide",
)

st.markdown("""
<style>
[data-testid="stAppViewContainer"]{background:#f8f9fb}
.main-header{background:linear-gradient(135deg,#1F3864 0%,#2E75B6 100%);padding:2rem 2.5rem;border-radius:16px;color:white;margin-bottom:2rem}
.main-header h1{font-size:1.8rem;font-weight:700;margin:0}
.main-header p{font-size:.95rem;opacity:.85;margin:.4rem 0 0}
.metric-box{background:#F0F4FF;border-radius:10px;padding:.9rem 1rem;border-left:4px solid #2E75B6;height:100%}
.metric-box .m-val{font-size:1.5rem;font-weight:700}
.metric-box .m-lbl{font-size:.75rem;color:#6B7A99;margin-top:2px}
.metric-box .m-sub{font-size:.7rem;color:#888;margin-top:2px}
.tag-done{background:#E6F4EA;color:#1E7E34;padding:2px 8px;border-radius:20px;font-size:.75rem;font-weight:600}
.tag-prog{background:#FFF8E1;color:#B45309;padding:2px 8px;border-radius:20px;font-size:.75rem;font-weight:600}
.tag-miss{background:#FDECEA;color:#B71C1C;padding:2px 8px;border-radius:20px;font-size:.75rem;font-weight:600}
.preview-table{width:100%;border-collapse:collapse;font-size:.85rem}
.preview-table th{background:#1F3864;color:white;padding:8px 12px;text-align:left}
.preview-table td{padding:7px 12px;border-bottom:1px solid #EEF0F4;vertical-align:top}
.preview-table tr:nth-child(even) td{background:#F8F9FF}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="main-header">
  <h1>📊 CS 결과보고 AI 자동작성</h1>
  <p>주간 현황판 + 개인 목표 파일을 올리면 AI가 OKR · KPI 포함 결과보고서를 자동으로 작성합니다.</p>
</div>
""", unsafe_allow_html=True)

# ── API 키
api_key = ""
try:
    api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
except Exception:
    pass
if not api_key:
    api_key = st.sidebar.text_input("Anthropic API Key", type="password")

# ── 사이드바
with st.sidebar:
    st.markdown("### ⚙️ 설정")
    author = st.text_input("작성자 이름", placeholder="예: 이호준")
    quarter = st.selectbox("보고 분기", [
        "2026년 2분기 (4~6월)",
        "2026년 3분기 (7~9월)",
        "2026년 1분기 (1~3월)",
        "2025년 4분기 (10~12월)",
    ])
    focus_hint = st.text_area("AI 분석 포커스 (선택)",
        placeholder="예: 계약갱신과 불만처리 중심으로\nKPI 달성률 강조", height=80)
    report_date = st.date_input("보고일", value=datetime.date.today())
    st.markdown("---")
    st.markdown("### 📤 출력 옵션")
    include_okr    = st.checkbox("OKR 요약 시트", value=True)
    include_kpi    = st.checkbox("KPI 실적 시트", value=True)
    include_detail = st.checkbox("날짜·상호·과정·결과 상세", value=True)
    include_eval   = st.checkbox("종합 평가 시트", value=True)
    apply_to_orig  = st.checkbox("원본 양식에 결과 주입", value=True)


# ═══════════════════════════════════════════════════════════════
# 파싱 — CS 주간 현황판 (4~6월 전체, 비고 포함 상세 추출)
# ═══════════════════════════════════════════════════════════════

def parse_cs_excel(file_bytes: bytes, person_name: str) -> dict:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # 작성자 시트 탐색 (이름 뒤 2글자, 전체 이름 순)
    target_sheet = None
    name_variants = [person_name, person_name[-2:], person_name[-3:]]
    for sname in wb.sheetnames:
        for nv in name_variants:
            if nv and nv in sname:
                target_sheet = sname
                break
        if target_sheet:
            break

    skip = {'원본시트(수정금지)', '----------', '📊 대시보드', '_옵션', '_성장노트 오류로그', '성장노트 종합'}
    sheets = [target_sheet] if target_sheet else [s for s in wb.sheetnames if not s.startswith('_') and s not in skip]

    result = {}
    for sname in sheets:
        if sname in skip or sname.startswith('_'):
            continue
        ws = wb[sname]
        weeks = []
        current_week = None
        in_table = False
        items = {}
        extra = {}   # 이번주 결과, 특이사항, 배운점

        ITEM_KEYS = {'점검','계약','미수','초과','오버홀','오버홀 ','해피콜','PC DB 및 홍보','IT기술력 습득','블로그 댓글'}
        Q2_MARKS  = ['26-04','26-05','26-06']

        after_result_header = False   # "목표 / 결과 / 목표" 헤더 다음 행 감지

        for row in ws.iter_rows(values_only=True):
            v0 = str(row[0]).strip() if row[0] else ''
            v3 = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            v6 = str(row[6]).strip() if len(row) > 6 and row[6] else ''

            # 새 주차 감지
            if '주간 현황판' in v0 and any(m in v0 for m in Q2_MARKS):
                if current_week:
                    weeks.append({'week': current_week, 'items': dict(items), 'extra': dict(extra)})
                current_week = v0; items = {}; extra = {}
                in_table = False; after_result_header = False

            if not current_week:
                continue

            if v0 == '항목':
                in_table = True
                continue

            if in_table and v0 in ITEM_KEYS:
                # 비고는 col[5], 갯수는 col[3] — 목표가 없으면 비고가 col[3]으로 밀릴 수 있음
                target_val = row[2] if len(row) > 2 else ''
                actual_val = row[3] if len(row) > 3 else ''
                rate_val   = row[4] if len(row) > 4 else ''
                note_val   = row[5] if len(row) > 5 else ''
                # 목표가 없고 비고가 col[3]에 있는 경우 (계약·점검 등)
                if not target_val and not actual_val and not note_val and v3:
                    note_val = v3
                items[v0] = {
                    'target': target_val,
                    'actual': actual_val,
                    'rate':   rate_val,
                    'note':   str(note_val or '')[:400],
                }
                continue

            if v0 == '이번주':
                in_table = False
                continue

            # "목표 / 결과 / 목표" 헤더 행 — 다음 행이 실제 결과
            if v0 == '목표' and v3 == '결과':
                after_result_header = True
                continue

            # 실제 결과 내용 행 (col[3] = 이번주 결과 텍스트)
            if after_result_header and v3 and len(v3) > 10:
                extra['결과'] = v3[:800]
                after_result_header = False

            # 특이사항
            if '특이사항' in v0:
                txt = str(row[0] or '') + ' ' + str(row[6] or '')
                extra['특이사항'] = txt[:400]

            # 배운점
            if '배운점' in v0:
                txt = str(row[0] or '') + ' ' + str(row[6] or '')
                extra['배운점'] = txt[:300]

        if current_week:
            weeks.append({'week': current_week, 'items': dict(items), 'extra': dict(extra)})

        if weeks:
            result[sname] = weeks

    return result


# ═══════════════════════════════════════════════════════════════
# 파싱 — 개인 목표
# ═══════════════════════════════════════════════════════════════

def parse_goal_excel(file_bytes: bytes) -> list:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    goals = []
    skip = {'_성장노트 오류로그', '성장노트 종합', '_옵션'}
    for sname in wb.sheetnames:
        if sname.startswith('_') or '수정금지' in sname or sname in skip:
            continue
        ws = wb[sname]
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx < 9:
                continue
            if not (row[1] and row[4]):
                continue
            goal_text = str(row[4]).strip()
            if len(goal_text) < 5:
                continue
            goals.append({
                'row':       row_idx,
                'sheet':     sname,
                'category':  str(row[1] or '').strip(),
                'grade':     str(row[3] or '').strip(),
                'goal':      goal_text[:250],
                'from_level': row[5],
                'to_level':  row[6],
                'mission':   str(row[10] or '').strip()[:150] if len(row) > 10 else '',
            })
    return goals


# ═══════════════════════════════════════════════════════════════
# CS 데이터를 AI용 압축 텍스트로 변환
# ═══════════════════════════════════════════════════════════════

def build_cs_text(cs_data: dict, max_chars: int = 5000) -> str:
    lines = []
    for sname, weeks in cs_data.items():
        for w in weeks:
            # 데이터가 없는 빈 주차는 건너뜀
            has_data = any(v.get('note') for v in w['items'].values()) or w['extra'].get('결과')
            if not has_data:
                continue
            lines.append(f"\n[{w['week']}]")
            for k, v in w['items'].items():
                note = str(v.get('note','')).replace('\n', ' ').strip()
                if not note:
                    continue
                target = v.get('target','') or '-'
                actual = v.get('actual','') or '-'
                lines.append(f"  {k}: 목표={target}, 실적={actual}, 비고={note[:200]}")
            # 이번주 결과 (가장 중요한 실제 실적 내용)
            if w['extra'].get('결과'):
                result_txt = w['extra']['결과'].replace('\n', ' / ').strip()
                lines.append(f"  ▶이번주결과: {result_txt[:400]}")
            if w['extra'].get('특이사항'):
                sp_txt = w['extra']['특이사항'].replace('\n', ' ').strip()
                lines.append(f"  ★특이사항: {sp_txt[:200]}")
    full = '\n'.join(lines)
    return full[:max_chars]


# ═══════════════════════════════════════════════════════════════
# AI 호출
# ═══════════════════════════════════════════════════════════════

def call_ai(cs_text: str, goal_text: str, author: str, quarter: str,
            focus: str, date_str: str, api_key: str) -> dict:
    client = anthropic.Anthropic(api_key=api_key)

    prompt = f"""당신은 CS 업무 결과보고 전문가입니다.
아래는 {author}의 {quarter} 실제 주간현황판 데이터입니다.
반드시 아래 데이터에 실제로 존재하는 날짜·상호명·수치·내용만 사용하세요. 임의로 만들지 마세요.
보고일: {date_str}
{f'분석 포커스: {focus}' if focus else ''}

## CS 주간 현황판 (실제 데이터)
{cs_text}

## 2분기 목표 목록
{goal_text}

위 실제 데이터를 바탕으로 아래 JSON 형식으로만 응답하세요 (```없이 순수 JSON):
{{
  "author": "{author}",
  "quarter": "{quarter}",
  "report_date": "{date_str}",
  "kpi_summary": [
    {{"item":"항목명","target":"주간목표수치","actual":"실제달성수치","rate":"달성율%","note":"주요내용(실제상호명포함)"}}
  ],
  "okr_summary": [
    {{"objective":"Objective 한줄","key_results":["KR1 실제수치포함","KR2","KR3"]}}
  ],
  "goal_results": [
    {{
      "category":"구분",
      "goal_title":"목표제목 30자이내",
      "okr_objective":"Objective 한줄",
      "key_results":["KR1 실제수치","KR2","KR3"],
      "details":[
        {{"date":"실제날짜(예:4/17)","company":"실제상호명","process":"실제 과정 2~3문장","result":"실제 결과"}}
      ],
      "achievement":"완료|진행중|미달",
      "kpi_rate":"달성율%"
    }}
  ],
  "cs_weekly_summary":[
    {{"item":"항목","target":"목표","actual":"실적","rate":"진행률","note":"실제내용"}}
  ],
  "overall_evaluation":{{
    "strengths":["실제근거있는 잘된점1","잘된점2","잘된점3"],
    "improvements":["아쉬운점1","아쉬운점2"],
    "next_quarter":["계획1","계획2","계획3"]
  }}
}}"""

    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=6000,
        messages=[{"role": "user", "content": prompt}]
    )
    raw = msg.content[0].text
    raw = re.sub(r'```json|```', '', raw).strip()
    m = re.search(r'\{[\s\S]*\}', raw)
    json_str = m.group() if m else raw

    try:
        return json.loads(json_str)
    except json.JSONDecodeError:
        fixed = json_str.rstrip().rstrip(',')
        if fixed.count('"') % 2 != 0:
            fixed += '"'
        open_arr = fixed.count('[') - fixed.count(']')
        open_obj = fixed.count('{') - fixed.count('}')
        fixed += ']' * max(open_arr, 0)
        fixed += '}' * max(open_obj, 0)
        try:
            return json.loads(fixed)
        except Exception:
            return {
                "author": author, "quarter": quarter, "report_date": date_str,
                "kpi_summary": [], "okr_summary": [],
                "goal_results": [], "cs_weekly_summary": [],
                "overall_evaluation": {
                    "strengths": ["AI 응답이 잘렸습니다. 재시도 해주세요."],
                    "improvements": [], "next_quarter": []
                }
            }


# ═══════════════════════════════════════════════════════════════
# 엑셀 생성
# ═══════════════════════════════════════════════════════════════

def thin_border():
    s = Side(style='thin', color='BFBFBF')
    return Border(top=s, bottom=s, left=s, right=s)

def hdr(ws, r, c, val, bg='1F3864', fg='FFFFFF', sz=10, bold=True, align='center'):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name='맑은 고딕', size=sz, bold=bold, color=fg)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    cell.border = thin_border()
    return cell

def dat(ws, r, c, val, bg='FFFFFF', fg='000000', sz=10, bold=False, align='left'):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font = Font(name='맑은 고딕', size=sz, bold=bold, color=fg)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical='top', wrap_text=True)
    cell.border = thin_border()
    return cell

def mhdr(ws, r, c1, c2, val, bg='1F3864', fg='FFFFFF', sz=11):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    return hdr(ws, r, c1, val, bg=bg, fg=fg, sz=sz)

def title_row(ws, r, c1, c2, val, author, quarter):
    ws.row_dimensions[r-1].height = 14
    ws.row_dimensions[r].height = 38
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    c = ws.cell(row=r, column=c1, value=val)
    c.font = Font(name='맑은 고딕', size=16, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='1F3864')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border()
    return r + 2


def build_result_excel(data: dict, author: str, quarter: str,
                       opts: dict, orig_goal_bytes) -> bytes:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── OKR 요약
    if opts.get('okr'):
        ws = wb.create_sheet('OKR 요약')
        ws.sheet_view.showGridLines = False
        for col, w in zip('ABCD', [3, 28, 55, 3]):
            ws.column_dimensions[col].width = w
        row = title_row(ws, 2, 2, 3, f'{author} · {quarter}  OKR 요약', author, quarter)
        for i, okr in enumerate(data.get('okr_summary', [])):
            ws.row_dimensions[row].height = 22
            mhdr(ws, row, 2, 3, f'Objective {i+1}:  {okr.get("objective","")}', bg='2E75B6', sz=11)
            row += 1
            for j, kr in enumerate(okr.get('key_results', []), 1):
                ws.row_dimensions[row].height = 20
                dat(ws, row, 2, f'KR{j}', bg='D6E4F7', fg='1F3864', bold=True, align='center')
                dat(ws, row, 3, kr)
                row += 1
            row += 1

    # ── KPI 실적
    if opts.get('kpi'):
        ws = wb.create_sheet('KPI 실적')
        ws.sheet_view.showGridLines = False
        for col, w in zip('ABCDEF', [3, 22, 14, 24, 10, 38]):
            ws.column_dimensions[col].width = w
        row = title_row(ws, 2, 2, 6, f'{author} · {quarter}  KPI 실적 현황', author, quarter)
        ws.row_dimensions[row].height = 22
        for ci, h in enumerate(['KPI 항목','목표','실적','달성율','비고']):
            hdr(ws, row, 2+ci, h)
        row += 1
        for i, r in enumerate(data.get('kpi_summary', [])):
            ws.row_dimensions[row].height = 22
            bg = 'FFFFFF' if i % 2 == 0 else 'F5F8FF'
            rate_str = str(r.get('rate',''))
            num = re.sub(r'[^0-9.]','', rate_str)
            rate_num = float(num) if num else 0
            fc = '375623' if rate_num >= 100 else ('C55A11' if rate_num >= 70 else 'C00000')
            dat(ws, row, 2, r.get('item',''),   bg=bg, bold=True, fg='1F3864')
            dat(ws, row, 3, r.get('target',''), bg=bg, align='center')
            dat(ws, row, 4, r.get('actual',''), bg=bg)
            dat(ws, row, 5, r.get('rate',''),   bg=bg, bold=True, fg=fc, align='center')
            dat(ws, row, 6, r.get('note',''),   bg=bg, sz=9)
            row += 1

    # ── 목표별 상세 결과 (날짜·상호·과정·결과)
    if opts.get('detail'):
        ws = wb.create_sheet('목표별 상세 결과')
        ws.sheet_view.showGridLines = False
        for col, w in zip('ABCDEFG', [3, 14, 11, 20, 46, 46, 3]):
            ws.column_dimensions[col].width = w
        row = title_row(ws, 2, 2, 6,
            f'{author} · {quarter}  목표별 상세 결과 (날짜 · 상호 · 과정 · 결과)', author, quarter)
        ws.row_dimensions[row].height = 22
        for ci, h in enumerate(['구분 / 목표','날짜','상호·항목','과정·실행 내용','결과·성과']):
            hdr(ws, row, 2+ci, h)
        row += 1

        COLOR_MAP = {
            '자기개발':  ('7030A0','EAD1DC'),
            '매출증대':  ('375623','E2EFDA'),
            '메출 증대': ('375623','E2EFDA'),
            '매출안정':  ('C55A11','FCE4D6'),
            '효율성':    ('2E75B6','D6E4F7'),
            '소통':      ('1F3864','BDD7EE'),
            '미션':      ('C00000','FFE7E7'),
        }

        for goal in data.get('goal_results', []):
            cat = goal.get('category','')
            hd_col, row_col = COLOR_MAP.get(cat, ('595959','F2F2F2'))
            details = goal.get('details', [])
            n = max(len(details), 1)
            sr = row

            ws.merge_cells(start_row=sr, start_column=2, end_row=sr+n-1, end_column=2)
            title_val = f'[{cat}]\n{goal.get("goal_title","")}\n▶ {goal.get("achievement","")}  {goal.get("kpi_rate","")}'
            c = ws.cell(row=sr, column=2, value=title_val)
            c.font = Font(name='맑은 고딕', size=9, bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor=hd_col)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin_border()

            for i, det in enumerate(details):
                r = row + i
                bg = 'FFFFFF' if i % 2 == 0 else row_col
                proc = str(det.get('process',''))
                res  = str(det.get('result',''))
                ws.row_dimensions[r].height = max(len(proc)//3, len(res)//3, 24)
                dat(ws, r, 3, det.get('date',''),    bg=bg, align='center', sz=9)
                dat(ws, r, 4, det.get('company',''), bg=bg, bold=True, sz=9)
                dat(ws, r, 5, proc,                  bg=bg, sz=9)
                dat(ws, r, 6, res,                   bg=bg, fg='375623', sz=9)

            if not details:
                ws.row_dimensions[row].height = 28
                for ci in range(3, 7):
                    dat(ws, row, ci, '-', bg=row_col, align='center')

            row += n
            ws.row_dimensions[row].height = 4
            row += 1

    # ── CS 주간 요약
    ws_cs = wb.create_sheet('CS 주간 요약')
    ws_cs.sheet_view.showGridLines = False
    for col, w in zip('ABCDEF', [3, 18, 11, 11, 10, 46]):
        ws_cs.column_dimensions[col].width = w
    row = title_row(ws_cs, 2, 2, 6, f'{author} · {quarter}  CS 주간 항목별 실적 요약', author, quarter)
    ws_cs.row_dimensions[row].height = 22
    for ci, h in enumerate(['항목','목표(주간)','실적(주간)','진행률','주요 내용']):
        hdr(ws_cs, row, 2+ci, h)
    row += 1
    for i, r in enumerate(data.get('cs_weekly_summary', [])):
        ws_cs.row_dimensions[row].height = 24
        bg = 'FFFFFF' if i % 2 == 0 else 'F5F8FF'
        dat(ws_cs, row, 2, r.get('item',''),   bg=bg, bold=True, fg='1F3864')
        dat(ws_cs, row, 3, r.get('target',''), bg=bg, align='center')
        dat(ws_cs, row, 4, r.get('actual',''), bg=bg, align='center')
        dat(ws_cs, row, 5, r.get('rate',''),   bg=bg, bold=True, align='center')
        dat(ws_cs, row, 6, r.get('note',''),   bg=bg, sz=9)
        row += 1

    # ── 종합 평가
    if opts.get('eval'):
        ws_ev = wb.create_sheet('종합 평가')
        ws_ev.sheet_view.showGridLines = False
        for col, w in zip('ABCD', [3, 16, 62, 3]):
            ws_ev.column_dimensions[col].width = w
        row = title_row(ws_ev, 2, 2, 3, f'{author} · {quarter}  종합 평가', author, quarter)
        ev = data.get('overall_evaluation', {})
        for section, items, bg_hd, bg_row in [
            ('✅ 잘된 점',       ev.get('strengths',[]),    '375623','E2EFDA'),
            ('⚠️ 아쉬운 점',    ev.get('improvements',[]), 'C55A11','FCE4D6'),
            ('📌 다음 분기 계획', ev.get('next_quarter',[]), '1F3864','D6E4F7'),
        ]:
            ws_ev.row_dimensions[row].height = 22
            mhdr(ws_ev, row, 2, 3, section, bg=bg_hd, sz=11)
            row += 1
            for item in items:
                ws_ev.row_dimensions[row].height = max(len(item)//5*3+18, 22)
                dat(ws_ev, row, 2, '•', bg=bg_row, align='center', bold=True)
                dat(ws_ev, row, 3, item, bg=bg_row)
                row += 1
            row += 1
        ws_ev.row_dimensions[row].height = 22
        ws_ev.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        dat(ws_ev, row, 2,
            f'작성자: {author}   |   보고일: {data.get("report_date","")}   |   결재: 팀장 □  부장 □  본부장 □',
            bg='F2F2F2', bold=True, align='center')

    # ── 원본 양식 주입
    if opts.get('orig') and orig_goal_bytes:
        try:
            wb_orig = openpyxl.load_workbook(io.BytesIO(orig_goal_bytes))
            for sname in wb_orig.sheetnames:
                if sname.startswith('_') or '수정금지' in sname:
                    continue
                ws_orig = wb_orig[sname]
                for row_idx in range(9, ws_orig.max_row + 1):
                    cat_cell  = ws_orig.cell(row=row_idx, column=2).value
                    if not cat_cell:
                        continue
                    cat_str = str(cat_cell).strip()
                    match = next((g for g in data.get('goal_results', [])
                                  if g.get('category','')[:4] in cat_str
                                  or cat_str[:4] in g.get('category','')), None)
                    if not match:
                        continue
                    detail_lines = '\n'.join(
                        f"▶ {d.get('date','')} {d.get('company','')} / {d.get('result','')}"
                        for d in match.get('details',[])[:3]
                    )
                    result_text = (
                        f"[{match.get('achievement','?')}] {match.get('kpi_rate','')} 달성\n"
                        f"{match.get('okr_objective','')}\n{detail_lines}"
                    )
                    c = ws_orig.cell(row=row_idx, column=10, value=result_text)
                    c.font = Font(name='맑은 고딕', size=9, color='0070C0')
                    c.alignment = Alignment(wrap_text=True, vertical='top')
            orig_buf = io.BytesIO()
            wb_orig.save(orig_buf)
            # 원본 주입본을 별도 시트로 메모
            ws_note = wb.create_sheet('원본양식_안내')
            ws_note['B2'] = '원본 양식 주입본은 아래 "원본양식_결과주입.xlsx" 로 별도 다운로드하세요.'
            ws_note['B2'].font = Font(color='C00000', bold=True)
        except Exception:
            pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════
# 메인 UI
# ═══════════════════════════════════════════════════════════════

col1, col2 = st.columns(2)
with col1:
    st.markdown("#### 📁 CS 주간 현황판")
    cs_file = st.file_uploader("CS 주간 현황판 Excel", type=['xlsx','xls'],
                                key='cs_file', label_visibility='collapsed')
    if cs_file:
        st.success(f"✅ {cs_file.name}")

with col2:
    st.markdown("#### 🎯 개인 2분기 목표")
    goal_file = st.file_uploader("개인 목표 Excel", type=['xlsx','xls'],
                                  key='goal_file', label_visibility='collapsed')
    if goal_file:
        st.success(f"✅ {goal_file.name}")

st.markdown("---")
ready = bool(cs_file and goal_file and author and api_key)
if not ready:
    missing = []
    if not cs_file:   missing.append("CS 주간 현황판")
    if not goal_file: missing.append("개인 목표 파일")
    if not author:    missing.append("작성자 이름")
    if not api_key:   missing.append("API Key")
    st.info(f"📋 필요: {' · '.join(missing)}")

run = st.button("🤖  AI 결과보고 자동 작성", use_container_width=True,
                 type="primary", disabled=not ready)

if run:
    cs_bytes   = cs_file.read()
    goal_bytes = goal_file.read()

    with st.status("AI가 데이터를 분석 중입니다...", expanded=True) as status:
        st.write("📂 파일 파싱 중...")
        cs_data   = parse_cs_excel(cs_bytes, author)
        goal_data = parse_goal_excel(goal_bytes)

        total_weeks = sum(len(v) for v in cs_data.values())
        st.write(f"✅ CS 파싱 완료 — {total_weeks}개 주차 · {len(goal_data)}개 목표")

        cs_text   = build_cs_text(cs_data, max_chars=4000)
        goal_lines = [
            f"{i+1}. [{g['category']}/{g['grade']}] {g['goal'][:200]}"
            + (f" / 미션: {g['mission'][:80]}" if g['mission'] else '')
            for i, g in enumerate(goal_data)
        ]
        goal_text = '\n'.join(goal_lines)[:2000]

        st.write("🤖 Claude 분석 중...")
        try:
            result = call_ai(cs_text, goal_text, author, quarter,
                             focus_hint, str(report_date), api_key)
            st.write(f"✅ 분석 완료 — 목표 {len(result.get('goal_results',[]))}개 · KPI {len(result.get('kpi_summary',[]))}개")
        except Exception as e:
            st.error(f"AI 오류: {e}")
            st.stop()

        st.write("📊 엑셀 생성 중...")
        opts = {
            'okr': include_okr, 'kpi': include_kpi,
            'detail': include_detail, 'eval': include_eval,
            'orig': apply_to_orig,
        }
        try:
            excel_bytes = build_result_excel(
                result, author, quarter, opts,
                goal_bytes if apply_to_orig else None
            )
            st.write("✅ 엑셀 생성 완료")
        except Exception as e:
            st.error(f"엑셀 오류: {e}")
            st.stop()

        status.update(label="✅ 결과보고 작성 완료!", state="complete")

    st.session_state['result']      = result
    st.session_state['excel_bytes'] = excel_bytes
    st.session_state['author']      = author
    st.session_state['quarter']     = quarter


# ── 결과 표시
if 'result' in st.session_state:
    result  = st.session_state['result']
    xl      = st.session_state['excel_bytes']
    _author = st.session_state['author']
    _qtr    = st.session_state['quarter']

    st.markdown("---")
    dl1, dl2 = st.columns([2, 1])
    with dl1:
        st.download_button(
            "📥 결과보고서 다운로드 (.xlsx)",
            data=xl,
            file_name=f"{_author}_{_qtr.split('(')[0].strip()}_결과보고.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary"
        )
    with dl2:
        st.download_button(
            "📋 JSON",
            data=json.dumps(result, ensure_ascii=False, indent=2).encode('utf-8'),
            file_name=f"{_author}_결과보고.json",
            mime="application/json",
            use_container_width=True
        )

    # KPI 카드
    kpi_list = result.get('kpi_summary', [])
    if kpi_list:
        st.markdown("#### 📈 KPI 달성 현황")
        cols = st.columns(min(len(kpi_list), 4))
        for i, kpi in enumerate(kpi_list[:8]):
            with cols[i % 4]:
                rate = kpi.get('rate','–')
                num  = re.sub(r'[^0-9.]','', str(rate))
                n    = float(num) if num else 0
                color = "#375623" if n >= 100 else ("#C55A11" if n >= 70 else "#C00000")
                st.markdown(f"""
                <div class="metric-box">
                  <div class="m-val" style="color:{color}">{rate}</div>
                  <div class="m-lbl">{kpi.get('item','')}</div>
                  <div class="m-sub">실적: {kpi.get('actual','')}</div>
                </div>""", unsafe_allow_html=True)

    # 목표별 요약 테이블
    goal_results = result.get('goal_results', [])
    if goal_results:
        st.markdown("---")
        st.markdown("#### 🎯 목표별 결과 요약")
        rows_html = ""
        for g in goal_results:
            ach = g.get('achievement','')
            cls = 'done' if ach=='완료' else ('prog' if ach=='진행중' else 'miss')
            krs = " / ".join(g.get('key_results',[])[:2])
            det_cnt = len(g.get('details',[]))
            rows_html += f"""<tr>
              <td><b style="color:#1F3864">{g.get('category','')}</b></td>
              <td><b>{g.get('goal_title','')}</b><br>
                  <span style="font-size:.78rem;color:#6B7A99">{g.get('okr_objective','')}</span></td>
              <td style="font-size:.78rem;color:#6B7A99">{krs}</td>
              <td><span class="tag-{cls}">{ach}</span></td>
              <td><b>{g.get('kpi_rate','–')}</b></td>
              <td style="text-align:center">{det_cnt}건</td>
            </tr>"""
        st.markdown(f"""
        <table class="preview-table">
          <thead><tr>
            <th>구분</th><th>목표</th><th>Key Results</th>
            <th>달성</th><th>KPI</th><th>상세</th>
          </tr></thead>
          <tbody>{rows_html}</tbody>
        </table>""", unsafe_allow_html=True)

    # CS 주간 요약
    cs_sum = result.get('cs_weekly_summary', [])
    if cs_sum:
        st.markdown("---")
        st.markdown("#### 📋 CS 항목별 요약")
        df = pd.DataFrame(cs_sum).rename(columns={
            'item':'항목','target':'목표','actual':'실적','rate':'진행률','note':'비고'})
        st.dataframe(df, use_container_width=True, hide_index=True)

    # 종합 평가
    ev = result.get('overall_evaluation', {})
    if ev:
        st.markdown("---")
        ec1, ec2, ec3 = st.columns(3)
        with ec1:
            st.markdown("**✅ 잘된 점**")
            for s in ev.get('strengths',[]): st.markdown(f"- {s}")
        with ec2:
            st.markdown("**⚠️ 아쉬운 점**")
            for s in ev.get('improvements',[]): st.markdown(f"- {s}")
        with ec3:
            st.markdown("**📌 다음 분기 계획**")
            for s in ev.get('next_quarter',[]): st.markdown(f"- {s}")
